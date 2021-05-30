# import openpyxl

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# # Load the workbook into openpyxl:
# wb = load_workbook('Grades.xlsx')

# # find out which is the active worksheet:
# ws = wb.active
# print(ws)

# # Access a specific sheet, not the active sheet:
# ws = wb['Sheet1']
# print(ws)

# # Change the value of cell A2 to the text 'Test':
# ws['A2'].value = 'Test'
# wb.save('Grades.xlsx') # You need to save the changes before it will take effect, also excel needs to be closed.

# # Name all the sheets in a workbook:
# print(wb.sheetnames)

# # Create a new sheet:
# wb.create_sheet("Test")
# wb.save('Grades.xlsx')
# print(wb.sheetnames)



# # Create a new workbook and manipulate it:
# wb = Workbook()
# ws = wb.active
# ws.title = "Data"

# ws.append(['Tim', 'is', 'great', '!'])

# wb.save('sam_test.xlsx')

# x = 0

# while x < 10:
#   ws.append(['Tim', 'is', 'great', '!'])
#   x += 1

# ws.append(["end"])
# wb.save('sam_test.xlsx')


# # Accessing multiple cells:
# wb = load_workbook('sam_test.xlsx')
# ws = wb.active

# for row in range(1, 15):
#   for col in range(1, 5):
#     char = get_column_letter(col)
#     # print(ws[char + str(row)])
#     print(ws[char + str(row)].value)
#     ws[char + str(row)] = char + str(row)

# wb.save('sam_test.xlsx')


# # Merging cells:
# wb = load_workbook('sam_test.xlsx')
# ws = wb.active

# ws.merge_cells("A1:D4")
# wb.save('sam_test.xlsx')

# ws.unmerge_cells("A1:D4")
# wb.save('sam_test.xlsx')

# Copy and move portions of a worksheet:
wb = load_workbook('sam_test.xlsx')
ws = wb.active

ws.insert_rows(7)
ws.insert_cols(7)

ws.move_range("C1:D11", rows=2, cols=2) 
# Moves C1:D22 2 rows down and 2 rows right. -ve nos. = up and left.

wb.save('sam_test.xlsx')