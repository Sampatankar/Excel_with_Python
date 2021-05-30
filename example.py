from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


data = {
  "Joe": {
    "math": 65,
    "science": 78,
    "english": 98,
    "gym": 89
  },
  "Bill": {
    "math": 55,
    "science": 72,
    "english": 87,
    "gym": 95
  },
  "Tim": {
    "math": 100,
    "science": 45,
    "english": 75,
    "gym": 92
  },
  "Sally": {
    "math": 30,
    "science": 25,
    "english": 45,
    "gym": 100
  },
  "Jane": {
    "math": 100,
    "science": 100,
    "english": 100,
    "gym": 60
  }
}

# Create workbook, and name the active worksheet 'Grades':
wb = Workbook()
ws = wb.active
ws.title = 'Grades'

"""
1. Create and object called 'heading's
2. The names need to be under a column named 'Names',
but all the other headings have keys already written so:
3. In headings, create a list called 'Names',
4. and create another list of all the keys.
We can iterate through the data structure that is under 'Joe'
and then use each key value to append to a list.
5. Finally heading is a concantenated list of 'Name' list and the other key value list.

"""
# Create the object 'heading' and concatenate lists:
headings = ['Name'] + list(data['Joe'].keys())
# Append to worksheet:
ws.append(headings)

"""
Loop through the dictionaries to get the data and append the
values and keys from the 'data' structure we created above,
under the headings we created above:
"""
for person in data:
  grades = list(data[person].values())
  ws.append([person] + grades)

"""
Loop through each column of interest and calculate the mean of their grades for that column.
We define where we place the mean, using range, such that if each person gets more or less
subjects they are graded for, (based on Joe's dictionary), it will change.
"""
# Where to get the data we want to base mean on:
for col in range(2, len(data['Joe']) + 2):
  # Get the column letter:
  char = get_column_letter(col)
  # Create the dynamic mean excel formula:
  ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

"""
Make all the headings bold, and change text colour.
Need to stylise every cell :/
"""
for col in range(1, 6):
  # Look at first row of each column, Bold and change to dark red font(hex):
  ws[get_column_letter(col) + '1'].font = Font(bold=True, color="8B0000")


# Save and execute the changes:
wb.save('NewGrades.xlsx')